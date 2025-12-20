import configparser
import inspect
import logging
import os
import time
import requests
from selenium import webdriver
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.chrome import options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
import pytest
from selenium.webdriver.chrome.options import Options
from configparser import ConfigParser
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from datetime import datetime
import allure
import json
from Utilities.configReader import readconfig
from Utilities.config import Config


# conftest_dir = os.path.dirname(os.path.abspath(__file__))
# config_filepath = os.path.join(conftest_dir, 'config.ini')
# config = configparser.ConfigParser(
#     # Pass 'None' to explicitly disable interpolation
#     interpolation=None)
#
# if not os.path.exists(config_filepath):  # Good practice to check if file exists
#     raise FileNotFoundError(f"Configuration file 'config.ini' not found at: {config_filepath}")
# try:
#     config.read(config_filepath)
#     print(f"DEBUG: Config loaded from: {config_filepath}")
#     print(f"DEBUG: Sections: {config.sections()}")
# except Exception as e:
#     raise RuntimeError(f"Error reading config file {config_filepath}: {e}")
#
# def readconfig(section, key):
#     try:
#         return config.get(section, key)
#     except configparser.NoSectionError:
#         # Provide helpful error message if section is missing
#         raise configparser.NoSectionError(f"No section '{section}' found in config file: {config_filepath}. Available sections: {config.sections()}")
#     except configparser.NoOptionError:
#         # Provide helpful error message if key is missing
#         raise configparser.NoOptionError(f"No option '{key}' found in section '{section}' in config file: {config_filepath}")

def pytest_addoption(parser):
    parser.addoption(
        "--browser_name", action="store", default="chrome", help="Specify browser: chrome, firefox, or edge"
    )

@pytest.fixture(scope="class")
def setup(request):
    driver = None  #
    browser_name = request.config.getoption("browser_name")
    print(f"Running tests on browser: {browser_name}")
    preferences = {"download.default_directory": readconfig("setup", "download_directory")}
    # # preferences = {"download.default_directory": (readconfig("setup", "download_directory"))}
    # ops = webdriver.ChromeOptions()
    # ops.add_argument("--ignore-certificate-errors")
    # ops.add_argument("--start-maximized")
    # ops.add_argument("--disable-notification")
    # # ops.add_argument("--incognito")  # to open in incognito mode
    # # ops.add_argument("headless")
    # ops.add_experimental_option("prefs", preferences)
    # driver = webdriver.Chrome(options=ops)

    # --------------
    browser_name = request.config.getoption("browser_name")
    if browser_name == "chrome":
        ops = webdriver.ChromeOptions()
        ops.add_argument("--ignore-certificate-errors")
        ops.add_argument("--start-maximized")
        ops.add_argument("--disable-notifications")
        ops.add_experimental_option("prefs", preferences)
        # ops.add_argument("headless") # to run driver in headless mode
        driver = webdriver.Chrome(options=ops)
    elif browser_name == "edge":
        ops = webdriver.EdgeOptions()
        ops.add_argument("--ignore-certificate-errors")
        # ops.add_argument("headless") # to run driver in headless mode
        ops.add_argument("--start-maximized")
        ops.add_argument("--disable-notifications")
        ops.add_experimental_option("prefs", preferences)
        driver = webdriver.Edge(options=ops)
    elif browser_name == "firefox":
        ops = webdriver.FirefoxOptions()
        ops.add_argument(
            "--ignore-certificate-errors")
        ops.add_argument("--start-maximized")
        ops.add_argument("--disable-notifications")
        ops.set_preference("browser.download.dir", readconfig("setup", "download_directory"))
        ops.set_preference("browser.download.folderList", 2)  # 0=desktop, 1=downloads, 2=custom location
        ops.set_preference("browser.download.useDownloadDir", True)
        ops.set_preference("browser.helperApps.neverAsk.saveToDisk",
                           "application/octet-stream")  # Common types to auto-download
        driver = webdriver.Firefox(options=ops)
    else:
        raise ValueError(f"Unsupported browser: {browser_name}. Please choose 'chrome', 'firefox', or 'edge'.")
    driver.implicitly_wait(20)
    # driver.get(readconfig("setup", "firstonline_url"))
    driver.get(Config.get("url"))
    request.cls.driver = driver
    yield driver
    print(f"Quitting {browser_name} browser.")
    driver.close()
    driver.quit()

def send_mail(sender_address, sender_pass, receiver_address, subject, mail_content, attach_file_name,
              ):
    # sender_pass = 'Selenium@234'

    # setting up the MIME
    message = MIMEMultipart()
    message['From'] = sender_address
    message['To'] = receiver_address
    message['Subject'] = subject

    # The subject line
    # The body and the attachments for the mail
    message.attach(MIMEText(mail_content, 'plain'))
    attach_file = open(attach_file_name, 'rb')  # Open the file as binary mode
    payload = MIMEBase('application', 'octate-stream')
    payload.set_payload((attach_file).read())
    encoders.encode_base64(payload)  # encode the attachment
    # add payload header with filename
    payload.add_header('Content-Disposition', 'attachment', filename=attach_file_name)
    message.attach(payload)

    # Create SMTP session for sending the mail
    session = smtplib.SMTP_SSL("smtp.gmail.com", 465)
    session.login(sender_address, sender_pass)  # login with mail_id and password
    text = message.as_string()
    session.sendmail(sender_address, receiver_address, text)
    session.quit()
    print('Mail Sent')

def read_data():
    file_path = readconfig("data", "file")
    sheet_name = readconfig("data", "sheet_name")
    print(f"Attempting to open file: '{file_path}'")
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
    except FileNotFoundError:
        pytest.fail(f"Excel file not found: {file_path}")
    except KeyError:
        pytest.fail(f"Sheet '{sheet_name}' not found in the Excel file")

    max_row = sheet.max_row
    max_col = sheet.max_column

    # Get headers from first row
    headers = [sheet.cell(row=1, column=col).value for col in range(1, max_col + 1)]
    print("headers are: ", headers)

    data_rows = []
    for row in range(2, max_row + 1):
        row_data = [sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
        data_rows.append(tuple(row_data))
    print(data_rows)
    return data_rows

def set_data(file_path, sheet_name, rowNum, colNum, data):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=rowNum, column=colNum).value = data
        workbook.save(file_path)

@pytest.hookimpl(tryfirst=True, hookwrapper=True)  # Correct decorator placement to embed failed screenshot to allure and html report
def pytest_runtest_makereport(item, call):
    pytest_html = item.config.pluginmanager.getplugin('html')
    outcome = yield # Capture the outcome of the test (pass/fail/skip)
    report = outcome.get_result()
    extra = getattr(report, 'extra', [])

    if report.when == 'call' or report.when == "setup": # Check if the failure happened during setup or call phase
        xfail = hasattr(report, 'wasxfail')
        if (report.skipped and xfail) or (report.failed and not xfail):
            # Attempt to get the driver from the test item's fixtures
            # This is the crucial part to access the driver instance that was used by the test
            driver = None
            try:
                # Access the driver from the 'setup' fixture.
                # 'item.funcargs' contains the arguments passed to the test function,
                # including fixtures.
                driver = item.funcargs.get('setup')
            except Exception as e:
                print(f"Could not retrieve driver from fixture: {e}")
                driver = None

            if driver:
                base_name = report.nodeid.split('[')[0]
                safe_name = base_name.replace("::", "__").replace("/", "_")
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

                # Define a directory for screenshots
                screenshot_dir = "reports/screenshots"
                os.makedirs(screenshot_dir, exist_ok=True)  # Create directory if it doesn't exist
                # Create a unique filename for the screenshot
                # Replace invalid characters for filenames
                file_name_base = f"{safe_name}_{timestamp}"
                file_name = os.path.join(screenshot_dir, f"{file_name_base}.png")

                try: #to attach screenshot to allure report
                    # 1. Save the screenshot to a file
                    driver.save_screenshot(file_name)
                    print(f"Screenshot saved: {file_name}")

                    # 2. Attach the screenshot file to the Allure report
                    with open(file_name, "rb") as image_file:
                        allure.attach(
                            image_file.read(),
                            name=f"Failure-screenshot - {timestamp}",
                            attachment_type=allure.attachment_type.PNG
                        )
                    print("Screenshot successfully attached to Allure report.")

                except Exception as e:
                    print(f"Failed to take or attach screenshot: {e}")

                try: # to attach screenshot to html report
                    driver.save_screenshot(file_name)  # Use driver.save_screenshot directly
                    print(f"Screenshot saved: {file_name}")

                    # Add the screenshot to the HTML report
                    if file_name:
                        # Construct the HTML for embedding the image in the report
                        html = f'<div><img src="{file_name}" alt="screenshot" style="width:304px;height:228px;" ' \
                               f'onclick="window.open(this.src)" align="right"/></div>'
                        extra.append(pytest_html.extras.html(html))
                except Exception as e:
                    print(f"Failed to take screenshot: {e}")
            else:
                print("Driver not available for screenshot.")
        report.extra = extra

@pytest.fixture(scope="session")
def test_data():
    """Load JSON test data once per test session."""
    # Get the directory where THIS conftest.py lives
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # Construct an absolute path to test_data.json
    file_path = os.path.join(base_dir, "test_data.json")

    # Debug print (optional)
    print(f"Loading JSON test data from: {file_path}")

    with open(file_path, "r") as f:
        data = json.load(f)

    if not isinstance(data, list):
        pytest.fail("JSON must contain a list of dictionaries")

    return data  # list of rows (dicts)

# @pytest.mark.hookwrapper
# @pytest.hookimpl(hookwrapper=True, tryfirst=True)
# def pytest_runtest_makereport(item, call):
#     outcome = yield
#     rep = outcome.get_result()
#     setattr(item, "rep_" + rep.when, rep)
#     return rep

