import inspect
import logging
import time
import requests
from selenium import webdriver
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.chrome import options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
import pytest
from selenium.webdriver.chrome.options import Options
from configparser import ConfigParser
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
import allure
import json
import openpyxl
import configparser
from pathlib import Path
import os



ROOT_DIR = Path(__file__).resolve().parent.parent
config_filepath = os.path.join(ROOT_DIR, 'config.ini')
config = configparser.ConfigParser(
    # Pass 'None' to explicitly disable interpolation
    interpolation=None)

if not os.path.exists(config_filepath):  # Good practice to check if file exists
    raise FileNotFoundError(f"Configuration file 'config.ini' not found at: {config_filepath}")
try:
    config.read(config_filepath)
    print(f"DEBUG: Config loaded from: {config_filepath}")
    print(f"DEBUG: Sections: {config.sections()}")
except Exception as e:
    raise RuntimeError(f"Error reading config file {config_filepath}: {e}")

def readconfig(section, key):
    try:
        return config.get(section, key)
    except configparser.NoSectionError:
        # Provide helpful error message if section is missing
        raise configparser.NoSectionError(f"No section '{section}' found in config file: {config_filepath}. Available sections: {config.sections()}")
    except configparser.NoOptionError:
        # Provide helpful error message if key is missing
        raise configparser.NoOptionError(f"No option '{key}' found in section '{section}' in config file: {config_filepath}")


excelfile = "//Users/mac/Desktop/PycharmProjects/InternetBanking/data/file.xlsx"
sheet_name = "details"
workbook = openpyxl.load_workbook(excelfile)
sheet = workbook[sheet_name]
totalrows =  sheet.max_row
totalcols = sheet.max_column

print("total rows are: ", str(totalrows), "and total column are: ", str(totalcols))

